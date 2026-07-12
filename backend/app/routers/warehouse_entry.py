from calendar import monthrange
from datetime import datetime

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse

from app.database import get_db
from app.middleware import require_role
from app.schemas.business import WarehouseEntryCreate

router = APIRouter(prefix="/api/v1/warehouse-entry", tags=["后工序入仓登记"])


def _build_send_pipeline(year: int | None, month: int | None,
                         product_code: str | None) -> list[dict]:
    """构建从后工序登记中提取送出记录的聚合管道。"""
    pipeline = []

    # 先筛选有送出记录的后工序文档
    pipeline.append({"$match": {
        "sends": {"$elemMatch": {
            "send_date": {"$ne": None},
            "send_qty": {"$gt": 0},
        }}
    }})

    # 展开 sends 数组，保留索引以生成唯一 ID
    pipeline.append({"$unwind": {"path": "$sends", "includeArrayIndex": "send_idx"}})

    # 过滤有效的送出记录
    pipeline.append({"$match": {
        "sends.send_date": {"$ne": None},
        "sends.send_qty": {"$gt": 0},
    }})

    # 月份筛选
    if year and month:
        last_day = monthrange(year, month)[1]
        start = datetime(year, month, 1)
        end = datetime(year, month, last_day)
        pipeline.append({"$match": {
            "sends.send_date": {"$gte": start, "$lte": end}
        }})

    # 产品编号筛选
    if product_code:
        pipeline.append({"$match": {
            "product_code": {"$regex": product_code, "$options": "i"}
        }})

    # 投影为入仓记录形状
    pipeline.append({"$project": {
        "_id": 0,
        "id": {"$concat": [
            {"$toString": "$_id"}, "_", {"$toString": "$send_idx"}
        ]},
        "entry_date": "$sends.send_date",
        "product_code": 1,
        "product_name": 1,
        "entry_qty": "$sends.send_qty",
        "material_code": {"$concat": [
            {"$dateToString": {"format": "%Y%m%d", "date": "$received_date"}},
            {"$substrCP": ["$shift", 0, 1]},
        ]},
    }})

    pipeline.append({"$sort": {"entry_date": -1}})
    return pipeline


@router.post("")
async def create_entry(data: WarehouseEntryCreate,
                       current=Depends(require_role("admin"))):
    db = get_db()
    record = {
        **data.model_dump(),
        "entry_date": datetime.combine(data.entry_date, datetime.min.time()),
        "created_at": datetime.utcnow(),
    }
    result = await db.warehouse_entry.insert_one(record)
    return {"message": "ok", "id": str(result.inserted_id)}


@router.get("")
async def list_entries(year: int | None = None,
                       month: int | None = None,
                       product_code: str | None = None,
                       page: int = 1, page_size: int = 500,
                       current=Depends(require_role("admin", "viewer"))):
    """查询入仓报表——从后工序登记中提取送出记录。"""
    db = get_db()
    pipeline = _build_send_pipeline(year, month, product_code)

    # 计数
    count_pipeline = [*pipeline, {"$count": "total"}]
    count_result = await db.post_process.aggregate(count_pipeline).to_list(length=1)
    total = count_result[0]["total"] if count_result else 0

    # 分页
    pipeline.append({"$skip": (page - 1) * page_size})
    pipeline.append({"$limit": page_size})

    records = await db.post_process.aggregate(pipeline).to_list(length=page_size)
    return {"total": total, "items": records, "page": page, "page_size": page_size}


@router.put("/{entry_id}")
async def update_entry(entry_id: str, data: WarehouseEntryCreate,
                       current=Depends(require_role("admin"))):
    db = get_db()
    result = await db.warehouse_entry.update_one(
        {"_id": ObjectId(entry_id)},
        {"$set": {
            **data.model_dump(),
            "entry_date": datetime.combine(data.entry_date, datetime.min.time()),
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="未找到该记录")
    return {"message": "ok"}


@router.delete("/{entry_id}")
async def delete_entry(entry_id: str,
                       current=Depends(require_role("admin"))):
    db = get_db()
    result = await db.warehouse_entry.delete_one({"_id": ObjectId(entry_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="未找到该记录")
    return {"message": "ok"}


@router.get("/export")
async def export_entries(year: int | None = None,
                         month: int | None = None,
                         product_code: str | None = None,
                         current=Depends(require_role("admin", "viewer"))):
    """导出入仓报表为 Excel。"""
    from io import BytesIO
    import openpyxl
    from openpyxl.utils import get_column_letter

    db = get_db()
    pipeline = _build_send_pipeline(year, month, product_code)
    pipeline.append({"$sort": {"entry_date": 1}})
    records = await db.post_process.aggregate(pipeline).to_list(length=10000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "后工序入仓报表"
    headers = ["入仓日期", "产品编号", "产品名称", "入仓数量", "物料编码"]
    ws.append(headers)
    for r in records:
        ws.append([
            r["entry_date"].strftime("%Y-%m-%d") if r.get("entry_date") else "",
            r.get("product_code", ""),
            r.get("product_name", ""),
            r.get("entry_qty", 0),
            r.get("material_code", ""),
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=warehouse_entry.xlsx"}
    )
