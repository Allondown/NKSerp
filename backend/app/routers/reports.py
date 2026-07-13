from datetime import datetime

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse

from app.database import get_db
from app.middleware import require_role

router = APIRouter(prefix="/api/v1/reports", tags=["统计报表"])


@router.get("/cost")
async def report_cost(year: int, month: int,
                      current=Depends(require_role("admin", "viewer"))):
    """月度成本报表：按材料规格统计期初/采购/领用/期末。"""
    db = get_db()
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)

    # 本月采购聚合
    purchase_pipeline = [
        {"$match": {"arrival_date": {"$gte": start_date, "$lt": end_date}}},
        {"$group": {
            "_id": "$material_spec",
            "purchase_qty": {"$sum": "$weight_kg"},
            "purchase_amount": {"$sum": "$total_price"}
        }}
    ]
    purchases = await db.purchase_records.aggregate(purchase_pipeline).to_list(length=500)

    # 本月领用聚合
    issue_pipeline = [
        {"$match": {"issue_date": {"$gte": start_date, "$lt": end_date}}},
        {"$group": {
            "_id": "$material_spec",
            "issue_qty": {"$sum": "$issue_weight_kg"},
            "issue_cost": {"$sum": "$total_cost"}
        }}
    ]
    issues = await db.issue_records.aggregate(issue_pipeline).to_list(length=500)

    p_map = {p["_id"]: p for p in purchases}
    i_map = {i["_id"]: i for i in issues}

    # 期初：当月之前所有采购和领用的累计
    prev_purchase_pipeline = [
        {"$match": {"arrival_date": {"$lt": start_date}}},
        {"$group": {
            "_id": "$material_spec",
            "prev_purchase_qty": {"$sum": "$weight_kg"},
            "prev_purchase_amount": {"$sum": "$total_price"}
        }}
    ]
    prev_purchases = await db.purchase_records.aggregate(prev_purchase_pipeline).to_list(length=500)
    prev_p_map = {p["_id"]: p for p in prev_purchases}

    prev_issue_pipeline = [
        {"$match": {"issue_date": {"$lt": start_date}}},
        {"$group": {
            "_id": "$material_spec",
            "prev_issue_qty": {"$sum": "$issue_weight_kg"},
            "prev_issue_cost": {"$sum": "$total_cost"}
        }}
    ]
    prev_issues = await db.issue_records.aggregate(prev_issue_pipeline).to_list(length=500)
    prev_i_map = {i["_id"]: i for i in prev_issues}

    # 仅显示当月有采购或领用记录的材料
    specs = set(list(p_map.keys()) + list(i_map.keys()))

    details = []
    total_purchase_amount = 0
    total_issue_cost = 0
    total_inventory_amount = 0

    for spec in sorted(specs):
        p = p_map.get(spec, {})
        i = i_map.get(spec, {})
        prev_p = prev_p_map.get(spec, {})
        prev_i = prev_i_map.get(spec, {})

        purchase_qty = p.get("purchase_qty", 0) or 0
        purchase_amount = p.get("purchase_amount", 0) or 0
        issue_qty = i.get("issue_qty", 0) or 0
        issue_cost = i.get("issue_cost", 0) or 0

        prev_purchase_qty = prev_p.get("prev_purchase_qty", 0) or 0
        prev_purchase_amount = prev_p.get("prev_purchase_amount", 0) or 0
        prev_issue_qty = prev_i.get("prev_issue_qty", 0) or 0
        prev_issue_cost = prev_i.get("prev_issue_cost", 0) or 0

        begin_qty = prev_purchase_qty - prev_issue_qty
        begin_amount = prev_purchase_amount - prev_issue_cost
        end_qty = begin_qty + purchase_qty - issue_qty
        end_amount = begin_amount + purchase_amount - issue_cost

        total_purchase_amount += purchase_amount
        total_issue_cost += issue_cost
        total_inventory_amount += end_amount

        details.append({
            "material_spec": spec,
            "begin_qty": round(begin_qty, 2), "begin_amount": round(begin_amount, 2),
            "purchase_qty": round(purchase_qty, 2), "purchase_amount": round(purchase_amount, 2),
            "issue_qty": round(issue_qty, 2), "issue_cost": round(issue_cost, 2),
            "end_qty": round(end_qty, 2), "end_amount": round(end_amount, 2),
        })

    return {
        "year": year, "month": month,
        "details": details,
        "summary": {
            "total_purchase_amount": round(total_purchase_amount, 2),
            "total_issue_cost": round(total_issue_cost, 2),
            "total_inventory_amount": round(total_inventory_amount, 2),
        }
    }


@router.get("/product-achieve")
async def report_product_achieve(year: int, month: int,
                                 current=Depends(require_role("admin", "viewer"))):
    """产品质量达成率报表：按产品+机器+班组。"""
    db = get_db()
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)

    pipeline = [
        {"$match": {"production_date": {"$gte": start_date, "$lt": end_date}}},
        {"$group": {
            "_id": {"product": "$product_name", "machine": "$machine"},
            "theoretical_qty": {"$sum": "$theoretical_qty"},
            "actual_qty": {"$sum": "$actual_qty"},
            "good_qty": {"$sum": "$good_qty"},
            "avg_qualified_rate": {"$avg": "$qualified_rate"},
        }},
        {"$sort": {"_id.product": 1, "_id.machine": 1}}
    ]
    rows = await db.daily_production.aggregate(pipeline).to_list(length=1000)

    result = []
    for r in rows:
        g = r["_id"]
        theo = r["theoretical_qty"] or 0
        actual = r["actual_qty"] or 0
        good = r["good_qty"] or 0
        result.append({
            "product": g["product"],
            "machine": g["machine"],
            "theoretical_qty": round(theo, 2),
            "actual_qty": actual,
            "good_qty": good,
            "bad_qty": max(actual - good, 0),
            "achieve_rate": round(actual / theo, 4) if theo > 0 else 0,
            "qualified_rate": round(r.get("avg_qualified_rate", 0), 4),
        })
    return result


@router.get("/team-achieve")
async def report_team_achieve(year: int, month: int,
                              current=Depends(require_role("admin", "viewer"))):
    """班组达成率汇总。"""
    db = get_db()
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)

    pipeline = [
        {"$match": {"production_date": {"$gte": start_date, "$lt": end_date}}},
        {"$group": {
            "_id": "$shift",
            "theoretical_qty": {"$sum": "$theoretical_qty"},
            "actual_qty": {"$sum": "$actual_qty"},
            "good_qty": {"$sum": "$good_qty"},
            "avg_qualified_rate": {"$avg": "$qualified_rate"},
        }}
    ]
    rows = await db.daily_production.aggregate(pipeline).to_list(length=10)
    result = []
    for r in rows:
        theo = r["theoretical_qty"] or 0
        actual = r["actual_qty"] or 0
        good = r["good_qty"] or 0
        result.append({
            "shift": r["_id"],
            "theoretical_qty": round(theo, 2),
            "actual_qty": actual,
            "good_qty": good,
            "achieve_rate": round(actual / theo, 4) if theo > 0 else 0,
            "qualified_rate": round(r.get("avg_qualified_rate", 0), 4),
        })
    return result


@router.get("/employee")
async def report_employee(year: int, month: int,
                          current=Depends(require_role("admin", "viewer"))):
    """员工绩效报表。"""
    db = get_db()
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)

    pipeline = [
        {"$match": {"production_date": {"$gte": start_date, "$lt": end_date}}},
        {"$group": {
            "_id": {"operator": "$operator", "product": "$product_name"},
            "theoretical_qty": {"$sum": "$theoretical_qty"},
            "actual_qty": {"$sum": "$actual_qty"},
            "good_qty": {"$sum": "$good_qty"},
            "avg_qualified_rate": {"$avg": "$qualified_rate"},
        }},
        {"$sort": {"_id.operator": 1}}
    ]
    rows = await db.daily_production.aggregate(pipeline).to_list(length=1000)

    result = []
    for r in rows:
        g = r["_id"]
        theo = r["theoretical_qty"] or 0
        actual = r["actual_qty"] or 0
        good = r["good_qty"] or 0
        result.append({
            "operator": g["operator"],
            "product": g["product"],
            "theoretical_qty": round(theo, 2),
            "actual_qty": actual,
            "good_qty": good,
            "achieve_rate": round(actual / theo, 4) if theo > 0 else 0,
            "qualified_rate": round(r.get("avg_qualified_rate", 0), 4),
        })
    return result


@router.get("/progress")
async def report_progress(year: int, month: int,
                          product_code: str | None = None,
                          current=Depends(require_role("admin", "viewer"))):
    """后工序完成进度报表：按产品+班组汇总送出数量。"""
    db = get_db()
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)

    match_filter = {"received_date": {"$gte": start_date, "$lt": end_date}}
    if product_code:
        match_filter["product_code"] = {"$regex": product_code, "$options": "i"}

    pipeline = [
        {"$match": match_filter},
        {"$addFields": {
            "total_send_qty": {
                "$reduce": {
                    "input": {"$ifNull": ["$sends", []]},
                    "initialValue": 0,
                    "in": {"$add": ["$$value", {"$ifNull": ["$$this.send_qty", 0]}]}
                }
            }
        }},
        {"$group": {
            "_id": {"product_code": "$product_code", "product_name": "$product_name"},
            "a_send": {"$sum": {"$cond": [{"$eq": ["$shift", "A班"]}, "$total_send_qty", 0]}},
            "b_send": {"$sum": {"$cond": [{"$eq": ["$shift", "B班"]}, "$total_send_qty", 0]}},
            "total_received": {"$sum": "$received_qty"},
        }},
        {"$sort": {"_id.product_code": 1}}
    ]
    rows = await db.post_process.aggregate(pipeline).to_list(length=500)

    result = []
    for r in rows:
        g = r["_id"]
        a = r["a_send"] or 0
        b = r["b_send"] or 0
        received = r["total_received"] or 0
        uncompleted = max(received - (a + b), 0)
        result.append({
            "product_code": g["product_code"],
            "product_name": g["product_name"],
            "a_total": a,
            "b_total": b,
            "ab_total": a + b,
            "uncompleted": uncompleted,
        })
    return result


@router.get("/post-process-summary")
async def report_post_process_summary(year: int,
                                       current=Depends(require_role("admin", "viewer"))):
    """后工序统计报表：机加送入月份未完成数量 + 后工序送出月份送出数量。"""
    db = get_db()

    # ---- 未完成统计：按送入月份汇总 ----
    uncompleted_pipeline = [
        {"$match": {
            "received_date": {
                "$gte": datetime(year, 1, 1),
                "$lte": datetime(year, 12, 31, 23, 59, 59),
            }
        }},
        {"$addFields": {
            "total_send_qty": {
                "$reduce": {
                    "input": {"$ifNull": ["$sends", []]},
                    "initialValue": 0,
                    "in": {"$add": ["$$value", {"$ifNull": ["$$this.send_qty", 0]}]}
                }
            }
        }},
        {"$group": {
            "_id": {
                "year": {"$year": "$received_date"},
                "month": {"$month": "$received_date"},
            },
            "total_received": {"$sum": "$received_qty"},
            "total_sent": {"$sum": "$total_send_qty"},
        }},
        {"$sort": {"_id.year": 1, "_id.month": 1}}
    ]
    uncompleted_rows = await db.post_process.aggregate(uncompleted_pipeline).to_list(length=1000)
    uncompleted = []
    for r in uncompleted_rows:
        g = r["_id"]
        received = r["total_received"] or 0
        sent = r["total_sent"] or 0
        uncompleted.append({
            "year": g["year"],
            "month": g["month"],
            "total_received": received,
            "total_sent": sent,
            "uncompleted": max(received - sent, 0),
        })

    # ---- 送出统计：按送出月份汇总 ----
    send_pipeline = [
        {"$unwind": {"path": "$sends", "preserveNullAndEmptyArrays": False}},
        {"$match": {
            "sends.send_date": {
                "$gte": datetime(year, 1, 1),
                "$lte": datetime(year, 12, 31, 23, 59, 59),
            }
        }},
        {"$group": {
            "_id": {
                "year": {"$year": "$sends.send_date"},
                "month": {"$month": "$sends.send_date"},
            },
            "total_send_qty": {"$sum": "$sends.send_qty"},
        }},
        {"$sort": {"_id.year": 1, "_id.month": 1}}
    ]
    send_rows = await db.post_process.aggregate(send_pipeline).to_list(length=1000)
    send_summary = []
    for r in send_rows:
        g = r["_id"]
        send_summary.append({
            "year": g["year"],
            "month": g["month"],
            "total_send_qty": r["total_send_qty"] or 0,
        })

    return {"uncompleted": uncompleted, "send_summary": send_summary}


@router.get("/tool-purchase-cost")
async def report_tool_purchase_cost(year: int, month: int,
                                    current=Depends(require_role("admin", "viewer"))):
    """刀具采购成本报表：按到货月份统计刀具采购成本。"""
    from calendar import monthrange
    db = get_db()
    last_day = monthrange(year, month)[1]
    start = datetime(year, month, 1)
    end = datetime(year, month, last_day, 23, 59, 59)

    records = await db.tool_purchases.find(
        {"arrival_date": {"$gte": start, "$lte": end}},
        sort=[("order_date", -1)]
    ).to_list(length=1000)

    items = []
    total_cost = 0.0
    for r in records:
        amount = r.get("total_amount", 0) or 0
        total_cost += amount
        items.append({
            "id": str(r.pop("_id")),
            "name": r.get("name", ""),
            "spec": r.get("spec", ""),
            "quantity": r.get("quantity", 0),
            "unit_price": r.get("unit_price", 0),
            "total_amount": amount,
            "processed_product": r.get("processed_product", ""),
            "supplier": r.get("supplier", ""),
            "order_date": r.get("order_date"),
            "arrival_date": r.get("arrival_date"),
        })
    return {"items": items, "total_cost": round(total_cost, 2)}


@router.get("/tool-supplier-cost")
async def report_tool_supplier_cost(year: int,
                                    current=Depends(require_role("admin", "viewer"))):
    """刀具采购供应商成本报表：按供应商和月份汇总采购成本。"""
    db = get_db()
    start = datetime(year, 1, 1)
    end = datetime(year, 12, 31, 23, 59, 59)

    pipeline = [
        {"$match": {"arrival_date": {"$gte": start, "$lte": end}}},
        {"$group": {
            "_id": {
                "supplier": "$supplier",
                "month": {"$month": "$arrival_date"},
            },
            "total_cost": {"$sum": "$total_amount"},
            "count": {"$sum": 1},
        }},
        {"$sort": {"_id.supplier": 1, "_id.month": 1}},
    ]
    rows = await db.tool_purchases.aggregate(pipeline).to_list(length=500)

    items = []
    grand_total = 0.0
    for r in rows:
        g = r["_id"]
        cost = round(r["total_cost"] or 0, 2)
        grand_total += cost
        items.append({
            "supplier": g["supplier"] or "",
            "month": g["month"],
            "total_cost": cost,
            "count": r["count"],
        })
    return {"items": items, "grand_total": round(grand_total, 2)}


@router.get("/export/excel")
async def export_excel(report_type: str, year: int, month: int,
                       product_code: str | None = None,
                       current=Depends(require_role("admin", "viewer"))):
    """导出报表为 Excel。"""
    from io import BytesIO
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{report_type}-{year}-{month}"

    if report_type == "cost":
        data = await report_cost(year, month)
        headers = ["材料规格", "期初数量", "期初金额", "采购数量", "采购金额",
                    "领用数量", "领用成本", "期末数量", "期末金额"]
        ws.append(headers)
        for d in data["details"]:
            ws.append([d["material_spec"], d["begin_qty"], d["begin_amount"],
                       d["purchase_qty"], d["purchase_amount"],
                       d["issue_qty"], d["issue_cost"],
                       d["end_qty"], d["end_amount"]])
        ws.append([])
        s = data["summary"]
        ws.append(["合计", "", "", "", s["total_purchase_amount"],
                    "", s["total_issue_cost"], "", s["total_inventory_amount"]])
    elif report_type == "product-achieve":
        data = await report_product_achieve(year, month)
        headers = ["产品", "机器", "理论产量", "实绩数量", "良品数量",
                    "不良数量", "达成率", "合格率"]
        ws.append(headers)
        for d in data:
            ws.append([d["product"], d["machine"],
                       d["theoretical_qty"], d["actual_qty"], d["good_qty"],
                       d["bad_qty"], d["achieve_rate"], d["qualified_rate"]])
    elif report_type == "team-achieve":
        data = await report_team_achieve(year, month)
        headers = ["班组", "理论产量", "实绩数量", "良品数量", "达成率", "合格率"]
        ws.append(headers)
        for d in data:
            ws.append([d["shift"], d["theoretical_qty"], d["actual_qty"],
                       d["good_qty"], d["achieve_rate"], d["qualified_rate"]])
    elif report_type == "employee":
        data = await report_employee(year, month)
        headers = ["操作工", "产品", "理论产量", "实绩数量", "良品数量", "达成率", "合格率"]
        ws.append(headers)
        for d in data:
            ws.append([d["operator"], d["product"], d["theoretical_qty"],
                       d["actual_qty"], d["good_qty"],
                       d["achieve_rate"], d["qualified_rate"]])
    elif report_type == "progress":
        data = await report_progress(year, month, product_code)
        headers = ["产品编号", "产品名称", "A班总数", "B班总数", "AB班完成总数", "未完成总数"]
        ws.append(headers)
        for d in data:
            ws.append([d["product_code"], d["product_name"],
                       d["a_total"], d["b_total"], d["ab_total"], d["uncompleted"]])
    elif report_type == "post-process-summary":
        data = await report_post_process_summary(year)
        headers = ["机加送入月份", "未完成数量"]
        ws.append(headers)
        total_uncompleted = 0
        for d in data["uncompleted"]:
            ws.append([f"{d['year']}-{d['month']:02d}", d["uncompleted"]])
            total_uncompleted += d["uncompleted"]
        ws.append(["合计", total_uncompleted])
        ws.append([])
        ws2 = wb.create_sheet("后工序送出统计")
        headers2 = ["后工序送出月份", "送出数量"]
        ws2.append(headers2)
        total_send = 0
        for d in data["send_summary"]:
            ws2.append([f"{d['year']}-{d['month']:02d}", d["total_send_qty"]])
            total_send += d["total_send_qty"]
        ws2.append(["合计", total_send])
        for col in range(1, len(headers2) + 1):
            ws2.column_dimensions[get_column_letter(col)].width = 16
    elif report_type == "tool-purchase-cost":
        data = await report_tool_purchase_cost(year, month)
        headers = ["品名", "规格", "数量", "单价", "总金额", "加工产品", "供应商", "下单日期", "到货日期"]
        ws.append(headers)
        for d in data["items"]:
            ws.append([
                d["name"], d["spec"], d["quantity"], d["unit_price"],
                d["total_amount"], d["processed_product"], d["supplier"],
                d["order_date"].strftime("%Y-%m-%d") if d.get("order_date") else "",
                d["arrival_date"].strftime("%Y-%m-%d") if d.get("arrival_date") else "",
            ])
        ws.append([])
        ws.append(["合计", "", "", "", data["total_cost"]])
    elif report_type == "tool-supplier-cost":
        data = await report_tool_supplier_cost(year)
        headers = ["供应商", "月份", "采购笔数", "采购金额"]
        ws.append(headers)
        for d in data["items"]:
            ws.append([d["supplier"], f"{d['month']}月", d["count"], d["total_cost"]])
        ws.append([])
        ws.append(["合计", "", "", data["grand_total"]])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{report_type}_{year}_{month}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
