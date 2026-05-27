from datetime import datetime

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse

from app.database import get_db
from app.middleware import require_role
from app.schemas.business import (DailyCombinedCreate, DailyProductionCreate,
                                   DailyProductionResponse, LossRemarkUpdate)

router = APIRouter(prefix="/api/v1/production", tags=["生产日报"])


@router.post("/daily")
async def create_daily(data: DailyProductionCreate,
                       current=Depends(require_role("admin", "workshop"))):
    db = get_db()
    theoretical_qty = data.work_time_sec / data.cycle_sec if data.cycle_sec > 0 else 0
    bad_qty = data.actual_qty - data.good_qty
    qualified_rate = data.good_qty / data.actual_qty if data.actual_qty > 0 else 0

    record = {
        **data.model_dump(),
        "production_date": datetime.combine(data.production_date, datetime.min.time()),
        "theoretical_qty": round(theoretical_qty, 4),
        "bad_qty": max(bad_qty, 0),
        "qualified_rate": round(qualified_rate, 4),
        "created_at": datetime.utcnow()
    }
    result = await db.daily_production.insert_one(record)
    return {"message": "ok", "id": str(result.inserted_id)}


@router.post("/daily/combined")
async def create_daily_combined(data: DailyCombinedCreate,
                                current=Depends(require_role("admin", "workshop"))):
    """AB班合并录入，自动拆分为两条记录。"""
    db = get_db()
    total = data.a_actual_qty + data.b_actual_qty
    fallback_a_work = round(data.work_time_sec * (data.a_actual_qty / total), 2) if total > 0 else round(data.work_time_sec / 2, 2)
    fallback_b_work = round(data.work_time_sec * (data.b_actual_qty / total), 2) if total > 0 else round(data.work_time_sec / 2, 2)
    a_work = data.a_work_time_sec if data.a_work_time_sec > 0 else fallback_a_work
    b_work = data.b_work_time_sec if data.b_work_time_sec > 0 else fallback_b_work
    a_cycle = data.a_cycle_sec if data.a_cycle_sec > 0 else data.cycle_sec
    b_cycle = data.b_cycle_sec if data.b_cycle_sec > 0 else data.cycle_sec

    def make_record(shift, operator, work_sec, cycle_sec, actual_qty, good_qty, loss_remark):
        theo = work_sec / cycle_sec if cycle_sec > 0 else 0
        bad = actual_qty - good_qty
        return {
            "production_date": datetime.combine(data.production_date, datetime.min.time()),
            "machine": data.machine,
            "product_name": data.product_name,
            "product_code": data.product_code,
            "material_spec": data.material_spec,
            "cycle_sec": cycle_sec,
            "work_time_sec": work_sec,
            "theoretical_qty": round(theo, 4),
            "good_qty": good_qty,
            "bad_qty": max(bad, 0),
            "actual_qty": actual_qty,
            "qualified_rate": round(good_qty / actual_qty, 4) if actual_qty > 0 else 0,
            "loss_time_min": data.loss_time_min,
            "shift": shift,
            "operator": operator,
            "plan_qty": data.plan_qty,
            "remark": data.remark,
            "loss_remark": loss_remark,
            "created_at": datetime.utcnow(),
        }

    records = []
    if data.a_actual_qty > 0 or data.a_operator:
        records.append(make_record("A班", data.a_operator, a_work, a_cycle, data.a_actual_qty, data.a_good_qty, data.a_loss_remark))
    if data.b_actual_qty > 0 or data.b_operator:
        records.append(make_record("B班", data.b_operator, b_work, b_cycle, data.b_actual_qty, data.b_good_qty, data.b_loss_remark))

    if not records:
        raise HTTPException(status_code=400, detail="请至少填写一个班组的数据")

    ids = []
    for rec in records:
        result = await db.daily_production.insert_one(rec)
        ids.append(str(result.inserted_id))
    return {"message": "ok", "ids": ids}


@router.put("/daily/combined")
async def update_daily_combined(
    original_date: str,
    original_machine: str,
    original_product_code: str,
    data: DailyCombinedCreate,
    current=Depends(require_role("admin", "workshop"))):
    """编辑合并日报：删除原记录并重新插入。"""
    db = get_db()
    orig_dt = datetime.strptime(original_date, "%Y-%m-%d")

    await db.daily_production.delete_many({
        "production_date": orig_dt,
        "machine": original_machine,
        "product_code": original_product_code,
    })

    total = data.a_actual_qty + data.b_actual_qty
    fallback_a_work = round(data.work_time_sec * (data.a_actual_qty / total), 2) if total > 0 else round(data.work_time_sec / 2, 2)
    fallback_b_work = round(data.work_time_sec * (data.b_actual_qty / total), 2) if total > 0 else round(data.work_time_sec / 2, 2)
    a_work = data.a_work_time_sec if data.a_work_time_sec > 0 else fallback_a_work
    b_work = data.b_work_time_sec if data.b_work_time_sec > 0 else fallback_b_work
    a_cycle = data.a_cycle_sec if data.a_cycle_sec > 0 else data.cycle_sec
    b_cycle = data.b_cycle_sec if data.b_cycle_sec > 0 else data.cycle_sec

    def make_record(shift, operator, work_sec, cycle_sec, actual_qty, good_qty, loss_remark):
        theo = work_sec / cycle_sec if cycle_sec > 0 else 0
        bad = actual_qty - good_qty
        return {
            "production_date": datetime.combine(data.production_date, datetime.min.time()),
            "machine": data.machine,
            "product_name": data.product_name,
            "product_code": data.product_code,
            "material_spec": data.material_spec,
            "cycle_sec": cycle_sec,
            "work_time_sec": work_sec,
            "theoretical_qty": round(theo, 4),
            "good_qty": good_qty,
            "bad_qty": max(bad, 0),
            "actual_qty": actual_qty,
            "qualified_rate": round(good_qty / actual_qty, 4) if actual_qty > 0 else 0,
            "loss_time_min": data.loss_time_min,
            "shift": shift,
            "operator": operator,
            "plan_qty": data.plan_qty,
            "remark": data.remark,
            "loss_remark": loss_remark,
            "created_at": datetime.utcnow(),
        }

    records = []
    if data.a_actual_qty > 0 or data.a_operator:
        records.append(make_record("A班", data.a_operator, a_work, a_cycle, data.a_actual_qty, data.a_good_qty, data.a_loss_remark))
    if data.b_actual_qty > 0 or data.b_operator:
        records.append(make_record("B班", data.b_operator, b_work, b_cycle, data.b_actual_qty, data.b_good_qty, data.b_loss_remark))

    if not records:
        raise HTTPException(status_code=400, detail="请至少填写一个班组的数据")

    ids = []
    for rec in records:
        result = await db.daily_production.insert_one(rec)
        ids.append(str(result.inserted_id))
    return {"message": "ok", "ids": ids}


@router.get("/daily/export")
async def export_daily(start_date: str | None = None,
                        end_date: str | None = None,
                        machine: str | None = None,
                        current=Depends(require_role("admin", "workshop", "warehouse", "viewer"))):
    """导出日报汇总为 Excel。"""
    from io import BytesIO
    import openpyxl
    from openpyxl.utils import get_column_letter

    db = get_db()
    match = {}
    if start_date or end_date:
        match["production_date"] = {}
        if start_date:
            match["production_date"]["$gte"] = datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            match["production_date"]["$lte"] = datetime.strptime(end_date, "%Y-%m-%d")
    if machine:
        match["machine"] = machine

    pipeline = [
        {"$match": match},
        {"$group": {
            "_id": {
                "date": "$production_date",
                "machine": "$machine",
                "product_code": "$product_code",
                "product_name": "$product_name",
            },
            "cycle_sec": {"$first": "$cycle_sec"},
            "plan_qty": {"$first": "$plan_qty"},
            "a_work_time": {"$sum": {"$cond": [{"$eq": ["$shift", "A班"]}, "$work_time_sec", 0]}},
            "b_work_time": {"$sum": {"$cond": [{"$eq": ["$shift", "B班"]}, "$work_time_sec", 0]}},
            "a_actual": {"$sum": {"$cond": [{"$eq": ["$shift", "A班"]}, "$actual_qty", 0]}},
            "b_actual": {"$sum": {"$cond": [{"$eq": ["$shift", "B班"]}, "$actual_qty", 0]}},
            "a_good": {"$sum": {"$cond": [{"$eq": ["$shift", "A班"]}, "$good_qty", 0]}},
            "b_good": {"$sum": {"$cond": [{"$eq": ["$shift", "B班"]}, "$good_qty", 0]}},
            "a_loss_remark": {"$max": {"$cond": [{"$eq": ["$shift", "A班"]}, "$loss_remark", ""]}},
            "b_loss_remark": {"$max": {"$cond": [{"$eq": ["$shift", "B班"]}, "$loss_remark", ""]}},
            "a_operator": {"$max": {"$cond": [{"$eq": ["$shift", "A班"]}, "$operator", ""]}},
            "b_operator": {"$max": {"$cond": [{"$eq": ["$shift", "B班"]}, "$operator", ""]}},
        }},
        {"$sort": {"_id.date": 1, "_id.machine": 1}}
    ]
    rows = await db.daily_production.aggregate(pipeline).to_list(length=10000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "日报汇总"
    headers = ["日期", "机器", "产品编号", "产品名称",
                "A节拍", "A生产时间", "A理论产量", "A实绩", "A良品", "A不良", "A合格率", "A损失备注", "A操作工",
                "B节拍", "B生产时间", "B理论产量", "B实绩", "B良品", "B不良", "B合格率", "B损失备注", "B操作工",
                "计划产量", "完成数量"]
    ws.append(headers)

    running_total = 0
    for r in rows:
        g = r["_id"]
        a_actual = r["a_actual"] or 0
        b_actual = r["b_actual"] or 0
        a_good = r["a_good"] or 0
        b_good = r["b_good"] or 0
        a_work = r["a_work_time"] or 0
        b_work = r["b_work_time"] or 0
        cycle = r["cycle_sec"] or 1
        a_theo = round(a_work / cycle, 2) if cycle > 0 else 0
        b_theo = round(b_work / cycle, 2) if cycle > 0 else 0
        running_total += a_good + b_good
        ws.append([
            g["date"].strftime("%Y-%m-%d") if g["date"] else "",
            g["machine"],
            g["product_code"],
            g["product_name"],
            cycle,
            a_work,
            a_theo,
            a_actual,
            a_good,
            max(a_actual - a_good, 0),
            f"{round(a_good / a_actual * 100, 1)}%" if a_actual > 0 else "0%",
            r.get("a_loss_remark", ""),
            r.get("a_operator", ""),
            cycle,
            b_work,
            b_theo,
            b_actual,
            b_good,
            max(b_actual - b_good, 0),
            f"{round(b_good / b_actual * 100, 1)}%" if b_actual > 0 else "0%",
            r.get("b_loss_remark", ""),
            r.get("b_operator", ""),
            r["plan_qty"] or 0,
            running_total,
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=daily_production.xlsx"}
    )


@router.get("/daily/export/list")
async def export_daily_list(start_date: str | None = None,
                             end_date: str | None = None,
                             machine: str | None = None,
                             product_code: str | None = None,
                             current=Depends(require_role("admin", "workshop", "warehouse", "viewer"))):
    """导出日报明细记录为 Excel（每次登记为唯一记录）。"""
    from io import BytesIO
    import openpyxl
    from openpyxl.utils import get_column_letter

    db = get_db()
    q = {}
    if start_date or end_date:
        q["production_date"] = {}
        if start_date:
            q["production_date"]["$gte"] = datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            q["production_date"]["$lte"] = datetime.strptime(end_date, "%Y-%m-%d")
    if machine:
        q["machine"] = machine
    if product_code:
        q["product_code"] = product_code

    records = await db.daily_production.find(q).sort("production_date", 1).to_list(length=10000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "日报明细"
    headers = ["日期", "机器", "产品编号", "产品名称", "节拍(秒/件)", "生产时间(秒)",
                "理论产量", "班组", "实绩数量", "良品数量", "不良数量", "合格率",
                "计划产量", "损失时间(分)", "操作工"]
    ws.append(headers)

    for r in records:
        actual = r.get("actual_qty", 0)
        good = r.get("good_qty", 0)
        qualified_rate = round(good / actual * 100, 1) if actual > 0 else 0
        theo = r.get("theoretical_qty", 0)
        ws.append([
            r["production_date"].strftime("%Y-%m-%d") if r.get("production_date") else "",
            r.get("machine", ""),
            r.get("product_code", ""),
            r.get("product_name", ""),
            r.get("cycle_sec", 0),
            r.get("work_time_sec", 0),
            theo,
            r.get("shift", ""),
            actual,
            good,
            max(actual - good, 0),
            f"{qualified_rate}%",
            r.get("plan_qty", 0),
            r.get("loss_time_min", 0),
            r.get("operator", ""),
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=daily_production_detail.xlsx"}
    )


@router.get("/daily/last/{product_code}")
async def last_daily(product_code: str,
                     current=Depends(require_role("admin", "workshop", "warehouse", "viewer"))):
    """查询指定产品最近一次日报记录，用于自动带出（含损失备注）。"""
    db = get_db()
    # 找最近有记录的两个日期，分别取 A/B 班数据
    latest = await db.daily_production.find_one(
        {"product_code": product_code},
        sort=[("production_date", -1), ("created_at", -1)]
    )
    if not latest:
        return None
    # 拿同一天的所有记录，提取 A/B 班损失备注
    same_day = await db.daily_production.find({
        "product_code": product_code,
        "production_date": latest["production_date"],
    }).to_list(length=10)
    a_loss_remark = ""
    b_loss_remark = ""
    for rec in same_day:
        if rec.get("shift") == "A班":
            a_loss_remark = rec.get("loss_remark", "") or ""
        elif rec.get("shift") == "B班":
            b_loss_remark = rec.get("loss_remark", "") or ""
    return {
        "product_name": latest.get("product_name", ""),
        "material_spec": latest.get("material_spec", ""),
        "machine": latest.get("machine", ""),
        "cycle_sec": latest.get("cycle_sec", 0),
        "plan_qty": latest.get("plan_qty", 0),
        "a_loss_remark": a_loss_remark,
        "b_loss_remark": b_loss_remark,
    }


@router.get("/daily")
async def list_daily(production_date: str | None = None,
                     machine: str | None = None,
                     shift: str | None = None,
                     product_code: str | None = None,
                     start_date: str | None = None,
                     end_date: str | None = None,
                     page: int = 1, page_size: int = 50,
                     current=Depends(require_role("admin", "workshop", "warehouse", "viewer"))):
    db = get_db()
    q = {}
    if production_date:
        q["production_date"] = datetime.strptime(production_date, "%Y-%m-%d")
    if machine:
        q["machine"] = machine
    if shift:
        q["shift"] = shift
    if product_code:
        q["product_code"] = product_code
    if start_date or end_date:
        q["production_date"] = {}
        if start_date:
            q["production_date"]["$gte"] = datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            q["production_date"]["$lte"] = datetime.strptime(end_date, "%Y-%m-%d")

    cursor = db.daily_production.find(q).sort("production_date", -1)
    total = await db.daily_production.count_documents(q)
    records = await cursor.skip((page - 1) * page_size).limit(page_size).to_list(length=page_size)
    for r in records:
        r["id"] = str(r.pop("_id"))
    return {"total": total, "items": records, "page": page, "page_size": page_size}


@router.get("/daily/summary")
async def daily_summary(start_date: str | None = None,
                        end_date: str | None = None,
                        machine: str | None = None,
                        current=Depends(require_role("admin", "workshop", "warehouse", "viewer"))):
    """日报汇总：按日期+机器+产品分组，AB班数据并排展示。"""
    db = get_db()
    match = {}
    if start_date or end_date:
        match["production_date"] = {}
        if start_date:
            match["production_date"]["$gte"] = datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            match["production_date"]["$lte"] = datetime.strptime(end_date, "%Y-%m-%d")
    if machine:
        match["machine"] = machine

    pipeline = [
        {"$match": match},
        {"$group": {
            "_id": {
                "date": "$production_date",
                "machine": "$machine",
                "product_code": "$product_code",
                "product_name": "$product_name",
            },
            "cycle_sec": {"$first": "$cycle_sec"},
            "plan_qty": {"$first": "$plan_qty"},
            "loss_time_min": {"$first": "$loss_time_min"},
            "a_work_time": {"$sum": {"$cond": [{"$eq": ["$shift", "A班"]}, "$work_time_sec", 0]}},
            "b_work_time": {"$sum": {"$cond": [{"$eq": ["$shift", "B班"]}, "$work_time_sec", 0]}},
            "a_actual": {"$sum": {"$cond": [{"$eq": ["$shift", "A班"]}, "$actual_qty", 0]}},
            "b_actual": {"$sum": {"$cond": [{"$eq": ["$shift", "B班"]}, "$actual_qty", 0]}},
            "a_good": {"$sum": {"$cond": [{"$eq": ["$shift", "A班"]}, "$good_qty", 0]}},
            "b_good": {"$sum": {"$cond": [{"$eq": ["$shift", "B班"]}, "$good_qty", 0]}},
            "a_loss_time_min": {"$max": {"$cond": [{"$eq": ["$shift", "A班"]}, "$loss_time_min", 0]}},
            "b_loss_time_min": {"$max": {"$cond": [{"$eq": ["$shift", "B班"]}, "$loss_time_min", 0]}},
            "a_loss_remark": {"$max": {"$cond": [{"$eq": ["$shift", "A班"]}, "$loss_remark", ""]}},
            "b_loss_remark": {"$max": {"$cond": [{"$eq": ["$shift", "B班"]}, "$loss_remark", ""]}},
            "a_operator": {"$max": {"$cond": [{"$eq": ["$shift", "A班"]}, "$operator", ""]}},
            "b_operator": {"$max": {"$cond": [{"$eq": ["$shift", "B班"]}, "$operator", ""]}},
        }},
        {"$sort": {"_id.date": -1, "_id.machine": 1}}
    ]
    rows = await db.daily_production.aggregate(pipeline).to_list(length=1000)

    result = []
    for r in rows:
        g = r["_id"]
        a_actual = r["a_actual"] or 0
        b_actual = r["b_actual"] or 0
        a_good = r["a_good"] or 0
        b_good = r["b_good"] or 0
        a_work = r["a_work_time"] or 0
        b_work = r["b_work_time"] or 0
        work_total = a_work + b_work
        cycle = r["cycle_sec"] or 1
        result.append({
            "production_date": g["date"].strftime("%Y-%m-%d") if g["date"] else "",
            "machine": g["machine"],
            "product_code": g["product_code"],
            "product_name": g["product_name"],
            "cycle_sec": cycle,
            "work_time_sec": work_total,
            "theoretical_qty": round(work_total / cycle, 2) if cycle > 0 else 0,
            "a_cycle_sec": cycle,
            "a_work_time_sec": a_work,
            "a_theoretical_qty": round(a_work / cycle, 2) if cycle > 0 else 0,
            "b_cycle_sec": cycle,
            "b_work_time_sec": b_work,
            "b_theoretical_qty": round(b_work / cycle, 2) if cycle > 0 else 0,
            "a_actual": a_actual,
            "a_good": a_good,
            "a_bad": max(a_actual - a_good, 0),
            "a_qualified_rate": round(a_good / a_actual, 4) if a_actual > 0 else 0,
            "b_actual": b_actual,
            "b_good": b_good,
            "b_bad": max(b_actual - b_good, 0),
            "b_qualified_rate": round(b_good / b_actual, 4) if b_actual > 0 else 0,
            "plan_qty": r["plan_qty"] or 0,
            "loss_time_min": r.get("loss_time_min", 0),
            "a_loss_time_min": r.get("a_loss_time_min", 0),
            "b_loss_time_min": r.get("b_loss_time_min", 0),
            "a_loss_remark": r.get("a_loss_remark", ""),
            "b_loss_remark": r.get("b_loss_remark", ""),
            "a_operator": r.get("a_operator", ""),
            "b_operator": r.get("b_operator", ""),
            "completion_qty": a_good + b_good,
        })
    return result


@router.put("/daily/update-loss-remark")
async def update_loss_remark(data: LossRemarkUpdate,
                             current=Depends(require_role("admin", "workshop"))):
    """更新指定班组记录的损失备注（内联编辑）。"""
    db = get_db()
    dt = datetime.strptime(data.production_date, "%Y-%m-%d")
    result = await db.daily_production.update_many(
        {
            "production_date": dt,
            "machine": data.machine,
            "product_code": data.product_code,
            "shift": data.shift,
        },
        {"$set": {"loss_remark": data.loss_remark}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="未找到匹配的记录")
    return {"message": "ok"}


@router.put("/daily/{record_id}")
async def update_daily(record_id: str, data: DailyProductionCreate,
                       current=Depends(require_role("admin", "workshop"))):
    db = get_db()
    theoretical_qty = data.work_time_sec / data.cycle_sec if data.cycle_sec > 0 else 0
    bad_qty = data.actual_qty - data.good_qty
    qualified_rate = data.good_qty / data.actual_qty if data.actual_qty > 0 else 0

    result = await db.daily_production.update_one(
        {"_id": ObjectId(record_id)},
        {"$set": {
            **data.model_dump(),
            "production_date": datetime.combine(data.production_date, datetime.min.time()),
            "theoretical_qty": round(theoretical_qty, 4),
            "bad_qty": max(bad_qty, 0),
            "qualified_rate": round(qualified_rate, 4),
        }}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="未找到该记录")
    return {"message": "ok"}


@router.delete("/daily/{record_id}")
async def delete_daily(record_id: str,
                       current=Depends(require_role("admin"))):
    db = get_db()
    result = await db.daily_production.delete_one({"_id": ObjectId(record_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="未找到该记录")
    return {"message": "ok"}
