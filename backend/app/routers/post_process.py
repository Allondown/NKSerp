from datetime import datetime

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse

from app.database import get_db
from app.middleware import require_role
from app.schemas.business import PostProcessCreate

router = APIRouter(prefix="/api/v1/post-process", tags=["后工序管理"])


def _convert_sends(sends: list) -> list:
    """Convert send entries: date → datetime."""
    result = []
    for s in sends:
        entry = {"send_qty": s.send_qty}
        if s.send_date:
            entry["send_date"] = datetime.combine(s.send_date, datetime.min.time())
        else:
            entry["send_date"] = None
        result.append(entry)
    return result


@router.post("")
async def create_record(data: PostProcessCreate,
                        current=Depends(require_role("admin", "workshop"))):
    db = get_db()
    record = {
        **data.model_dump(exclude={"sends"}),
        "received_date": datetime.combine(data.received_date, datetime.min.time()),
        "sends": _convert_sends(data.sends),
        "created_at": datetime.utcnow()
    }
    result = await db.post_process.insert_one(record)
    return {"message": "ok", "id": str(result.inserted_id)}


@router.get("")
async def list_records(start_date: str | None = None,
                       end_date: str | None = None,
                       product_code: str | None = None,
                       page: int = 1, page_size: int = 100,
                       current=Depends(require_role("admin", "workshop", "warehouse", "viewer"))):
    db = get_db()
    q = {}
    if start_date or end_date:
        q["received_date"] = {}
        if start_date:
            q["received_date"]["$gte"] = datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            q["received_date"]["$lte"] = datetime.strptime(end_date, "%Y-%m-%d")
    if product_code:
        q["product_code"] = {"$regex": product_code, "$options": "i"}

    cursor = db.post_process.find(q).sort("received_date", -1)
    total = await db.post_process.count_documents(q)
    records = await cursor.skip((page - 1) * page_size).limit(page_size).to_list(length=page_size)
    for r in records:
        r["id"] = str(r.pop("_id"))
        sends = r.get("sends") or []
        r["total_send_qty"] = sum(s.get("send_qty", 0) or 0 for s in sends)
    return {"total": total, "items": records, "page": page, "page_size": page_size}


@router.put("/{record_id}")
async def update_record(record_id: str, data: PostProcessCreate,
                        current=Depends(require_role("admin", "workshop"))):
    db = get_db()
    update_data = {
        **data.model_dump(exclude={"sends"}),
        "received_date": datetime.combine(data.received_date, datetime.min.time()),
        "sends": _convert_sends(data.sends),
    }

    result = await db.post_process.update_one(
        {"_id": ObjectId(record_id)},
        {"$set": update_data}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="未找到该记录")
    return {"message": "ok"}


@router.get("/export")
async def export_records(start_date: str | None = None,
                          end_date: str | None = None,
                          product_code: str | None = None,
                          current=Depends(require_role("admin", "workshop", "warehouse", "viewer"))):
    """导出后工序登记记录为 Excel。"""
    from io import BytesIO
    import openpyxl
    from openpyxl.utils import get_column_letter

    db = get_db()
    q = {}
    if start_date or end_date:
        q["received_date"] = {}
        if start_date:
            q["received_date"]["$gte"] = datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            q["received_date"]["$lte"] = datetime.strptime(end_date, "%Y-%m-%d")
    if product_code:
        q["product_code"] = {"$regex": product_code, "$options": "i"}

    records = await db.post_process.find(q).sort("received_date", -1).to_list(length=10000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "后工序登记"
    headers = ["机加送入日期", "产品编号", "产品名称", "送入数量",
                "操作员", "班组", "送出序号", "送出日期", "送出数量", "未完成数量"]
    ws.append(headers)

    for r in records:
        received = r.get("received_qty", 0) or 0
        sends = r.get("sends") or []
        total_sent = sum(s.get("send_qty", 0) or 0 for s in sends)
        if not sends:
            ws.append([
                r.get("received_date").strftime("%Y-%m-%d") if r.get("received_date") else "",
                r.get("product_code", ""),
                r.get("product_name", ""),
                received,
                r.get("operator", ""),
                r.get("shift", ""),
                "-", "", 0,
                max(received - total_sent, 0),
            ])
        else:
            # 找出最新送出日期的索引，只在该行显示未完成数量
            def _date_key(idx_s):
                d = idx_s[1].get("send_date")
                return d if d else datetime.min
            latest_idx = max(enumerate(sends), key=_date_key)[0]
            for i, s in enumerate(sends, 1):
                remaining = max(received - total_sent, 0) if (i - 1) == latest_idx else ""
                ws.append([
                    r.get("received_date").strftime("%Y-%m-%d") if r.get("received_date") else "",
                    r.get("product_code", ""),
                    r.get("product_name", ""),
                    received,
                    r.get("operator", ""),
                    r.get("shift", ""),
                    f"S{i}",
                    s.get("send_date").strftime("%Y-%m-%d") if s.get("send_date") else "",
                    s.get("send_qty", 0) or 0,
                    remaining,
                ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=post_process.xlsx"}
    )


@router.delete("/{record_id}")
async def delete_record(record_id: str,
                        current=Depends(require_role("admin"))):
    db = get_db()
    result = await db.post_process.delete_one({"_id": ObjectId(record_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="未找到该记录")
    return {"message": "ok"}
